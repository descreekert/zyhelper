"""
从 2026高考-物理-志愿_2026.xlsx 导出 JSON, 给 web app 用.

输出:
    data/plans.json          - 招生计划数据 (12728 条)
    data/score_rank.json     - 一分一段表 (2022-2026) + 等位分对照
    data/priority.json       - 城市/学校/专业类 三张优先次序表 (来自 排名 sheet)
    data/meta.json           - 数据生成时间、来源、枚举值

依赖: openpyxl
"""

import csv
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl

SRC = Path("/Users/banlibo/Desktop/高考/new/2026高考-物理-志愿_2026.xlsx")
OUT_DIR = Path(__file__).resolve().parent.parent / "data"

# === V9: 专业保研率 新算法数据源 ===
# 算法 (per major):
#   1. 在 school_majors_data 找 (校, 专) -> 学院
#      兜底: 院校-学院-专业保研率.csv 也含 (校, 专, 学院) 映射
#   2. base_rate = 校保研率 (plan.schoolBaoyan)
#   3. override_rate = 院校-学院保研率.csv 的 (校, 学院) 命中
#   4. rate = override if 命中 else base
#   5. 输出 "{专业}({学院},{rate%})"
SCHOOL_MAJORS_DATA = Path("/Users/banlibo/Desktop/高考/old/院校-学院-专业/school_majors_data")
COLLEGE_BAOYAN_CSV = Path("/Users/banlibo/Desktop/高考/new/工具/保研率/院校-学院保研率.csv")
MAJOR_BAOYAN_CSV   = Path("/Users/banlibo/Desktop/高考/new/工具/保研率/院校-学院-专业保研率.csv")  # 也用作 (校,专)->学院 fallback

# ----- 工具 -----
def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_str(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_contained_majors(s):
    """所含专业可能是 '逗号' 或 '、' 分隔."""
    if not s:
        return []
    parts = re.split(r"[,、]", str(s))
    return [p.strip() for p in parts if p.strip()]


# ----- 变化字段 diff 解析 (数据-2) -----
def _split_prefix_majors(s):
    """切分 'prefix(majors)' 或 'prefix[majors]', 返回 (prefix, majors_list).

    majors 是最后一个 () 或 [] 内的 逗号/顿号 分隔列表.
    """
    s = (s or "").strip()
    if not s:
        return s, []
    if s[-1] in (")", "]", "）", "】"):
        # 找匹配开括号
        close_to_open = {")": "(", "]": "[", "）": "（", "】": "【"}
        open_ch = close_to_open[s[-1]]
        close_ch = s[-1]
        depth = 0
        last_open = -1
        for i in range(len(s) - 1, -1, -1):
            c = s[i]
            if c == close_ch:
                depth += 1
            elif c == open_ch:
                depth -= 1
                if depth == 0:
                    last_open = i
                    break
        if last_open > 0:
            prefix = s[:last_open]
            inner = s[last_open + 1:-1]
            majors = [m.strip() for m in re.split(r"[、,，]", inner) if m.strip()]
            return prefix, majors
    return s, []


def parse_diff(diff_str):
    """解析"变化"字段, 返回结构化 list.

    单段格式:
      - 招生专业改名: X → Y
      - 招生人数 +N / -N
      - 学费 X → Y
      - 新增: X
      - 停招: X

    多段以分号 ; 分隔.
    """
    if not diff_str:
        return []
    parts = [p.strip() for p in re.split(r"[;；]", str(diff_str)) if p.strip()]
    results = []
    for part in parts:
        # 招生专业改名
        m = re.match(r"招生专业改名[：:]\s*(.*?)\s*→\s*(.*)", part)
        if m:
            old_full, new_full = m.group(1), m.group(2)
            old_prefix, old_majors = _split_prefix_majors(old_full)
            new_prefix, new_majors = _split_prefix_majors(new_full)
            old_set, new_set = set(old_majors), set(new_majors)
            results.append({
                "type": "rename",
                "oldPrefix": old_prefix,
                "newPrefix": new_prefix,
                "prefixSame": old_prefix == new_prefix,
                "added":   sorted(new_set - old_set),
                "removed": sorted(old_set - new_set),
                "kept":    sorted(new_set & old_set),
            })
            continue
        # 招生人数
        m = re.match(r"招生人数([+\-]\d+)", part)
        if m:
            results.append({"type": "num", "delta": int(m.group(1))})
            continue
        # 学费
        m = re.match(r"学费\s*(\d+)\s*→\s*(\d+)", part)
        if m:
            results.append({
                "type": "tuition", "from": int(m.group(1)), "to": int(m.group(2)),
            })
            continue
        # 新增
        m = re.match(r"新增[：:]\s*(.*)", part)
        if m:
            results.append({"type": "new", "text": m.group(1).strip()})
            continue
        # 停招
        m = re.match(r"停招[：:]\s*(.*)", part)
        if m:
            results.append({"type": "stopped", "text": m.group(1).strip()})
            continue
        # 其它
        results.append({"type": "other", "text": part})
    return results


def diff_summary(structured):
    """给 list of structured diff, 生成短摘要 (供表格列展示).

    e.g. '改名 +3-7, 人数-5'
    """
    if not structured:
        return ""
    bits = []
    for d in structured:
        t = d["type"]
        if t == "rename":
            a, r = len(d["added"]), len(d["removed"])
            tag = "改名"
            if d.get("prefixSame") is False:
                tag = "改名(类名也改)"
            bits.append(f"{tag} +{a}-{r}")
        elif t == "num":
            bits.append(f"人数{d['delta']:+d}")
        elif t == "tuition":
            bits.append(f"学费 {d['from']}→{d['to']}")
        elif t == "new":
            bits.append("新增")
        elif t == "stopped":
            bits.append("停招")
        else:
            bits.append(d.get("text", "")[:20])
    return ", ".join(bits)


# ----- 主数据 (plans) -----
def load_plans(wb):
    ws = wb["2026数据raw"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(hdr) if h}

    def g(row, col_name):
        i = idx.get(col_name)
        return row[i] if i is not None else None

    plans = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        sc_code = to_str(g(row, "院校编号"))
        if not sc_code:
            continue

        # 历年录取
        history = {}
        for yr in [24, 23, 22]:
            num = to_int(g(row, f"{yr}人数"))
            score = to_int(g(row, f"{yr}最低分"))
            rank = to_int(g(row, f"{yr}最低位次"))
            line_diff = to_int(g(row, f"{yr}线差"))
            if num is None and score is None:
                history[str(yr)] = None
            else:
                history[str(yr)] = {
                    "num": num,
                    "score": score,
                    "rank": rank,
                    "lineDiff": line_diff,
                }

        # 中外合作办学判别
        remarks = to_str(g(row, "专业备注"))
        major_name_26 = to_str(g(row, "招生专业"))
        is_mid_outside = ("中外合作办学" in remarks) or ("中外合作办学" in major_name_26)

        # 停招判别 (25参考可信度 为空)
        ref_conf = to_str(g(row, "25参考可信度"))
        is_stopped = (ref_conf == "")

        plan = {
            "id": to_int(g(row, "序号")),
            # 学校
            "schoolCode":   sc_code,
            "schoolName":   to_str(g(row, "招生院校")),
            "schoolRank":   to_int(g(row, "学校排名")),
            "schoolTag":    to_str(g(row, "院校标签")),
            "schoolType":   to_str(g(row, "类型")),
            "province":     to_str(g(row, "所在省")),
            "city":         to_str(g(row, "城市")),
            "cityTier":     to_str(g(row, "城市水平")),
            "cityPriority": to_int(g(row, "城市优先")),
            "schoolPriority": to_int(g(row, "学校优先")),
            "majorPriority": to_int(g(row, "专业类优先")),
            "schoolBaoyan": to_float(g(row, "校保研率")),
            "schoolUpgrade": to_float(g(row, "校升学率")),
            "managing":     to_str(g(row, "主管部门")),
            "hasBaoyan":    to_str(g(row, "保研资格")),

            # 26
            "majorCode26":     to_str(g(row, "专业编号")),
            "majorCategory":   to_str(g(row, "门类")),
            "majorClass":      to_str(g(row, "专业类")),
            "majorName26":     major_name_26,
            "containedMajors": parse_contained_majors(g(row, "所含专业")),
            "rankSoftware":    to_str(g(row, "软科专业排名")),
            "rankEval":        to_str(g(row, "学科评估")),
            "rankMaster":      to_str(g(row, "软科硕士学科排名")),
            "baoyanDetail":    to_str(g(row, "专业保研率")),
            "remarks":         remarks,
            "subjectReq":      to_str(g(row, "选科要求")),
            "duration":        to_str(g(row, "学制")),
            "tuition":         to_int(g(row, "学费")) if to_int(g(row, "学费")) is not None
                                                     else to_str(g(row, "学费")),
            "isNew":           to_str(g(row, "是否新增")),
            "enrollNum26":     to_int(g(row, "人数")),
            "diff":            to_str(g(row, "变化")),
            "diffStructured":  parse_diff(to_str(g(row, "变化"))),
            "diffSummary":     diff_summary(parse_diff(to_str(g(row, "变化")))),

            # 25 参考
            "ref25Score":   to_int(g(row, "25参考最低分")),
            "ref25Rank":    to_int(g(row, "25参考最低位次")),
            "ref25LineDiff": to_float(g(row, "25参考线差")),
            "refConfidence": ref_conf,
            "refSource":    to_str(g(row, "25参考来源")),

            # 25 实际
            "majorName25":   to_str(g(row, "25招生专业")),
            "majorClass25":  to_str(g(row, "25专业类")),
            "rankSoftware25": to_str(g(row, "25软科专业排名")),
            "rankEval25":    to_str(g(row, "25学科评估")),
            "rankMaster25":  to_str(g(row, "25软科硕士学科排名")),
            "baoyanDetail25": to_str(g(row, "25专业保研率")),
            "remarks25":     to_str(g(row, "25分流转专业")),
            "subjectReq25":  to_str(g(row, "25选科要求")),
            "duration25":    to_str(g(row, "25学制")),
            "tuition25":     to_int(g(row, "25学费")) if to_int(g(row, "25学费")) is not None
                                                       else to_str(g(row, "25学费")),
            "isNew25":       to_str(g(row, "25是否新增")),
            "enrollNum25":   to_int(g(row, "25人数")),
            "score25":       to_int(g(row, "25最低分")),
            "rank25":        to_int(g(row, "25最低位次")),
            "lineDiff25":    to_int(g(row, "25线差")),

            # 历年
            "history": history,

            # 统计 / 预测
            "avgRank":       to_int(g(row, "平均位次")),
            "avgLineDiff":   to_int(g(row, "平均线差")),
            "rankVolatility": to_float(g(row, "25位次波动")),
            "predict": {
                "num":       to_int(g(row, "预测人数")),
                "score":     to_int(g(row, "预测分数")),
                "rank":      to_int(g(row, "预测位次")),
                "lineDiff":  to_int(g(row, "预测线差")),
                "lineScore": to_int(g(row, "线差预测分数")),
                "heat":      to_float(g(row, "预测热度")),
                "rankShift": to_float(g(row, "预测位次波动")),
                "trend":     to_str(g(row, "预测趋势")),
            },

            # 标记 (预计算)
            "isStopped":     is_stopped,
            "isMidOutside":  is_mid_outside,
        }
        plans.append(plan)
    return plans


# ----- 一分一段 + 等位分 -----
def load_score_rank(wb):
    ws = wb["分数"]
    rows = list(ws.iter_rows(values_only=True))

    # 一本线 (年份, 一本线)
    baseline = {}
    for r in rows[1:8]:   # row 2..7 是年份数据
        year = to_int(r[0])
        ben = to_int(r[2])
        if year and ben:
            baseline[str(year)] = ben

    # 一分一段表的列位置:
    #   2026: col 30-32  (分数, 本分人数, 累计人数)
    #   2025: col 34-36
    #   2024: col 38-40
    #   2023: col 42-44
    #   2022: col 46-48
    one_score = {}
    year_cols = {
        "2026": 30 - 1,
        "2025": 34 - 1,
        "2024": 38 - 1,
        "2023": 42 - 1,
        "2022": 46 - 1,
    }
    for year, col0 in year_cols.items():
        table = []
        for r in rows[2:]:   # row 3 起是数据 (row 1=标题, row 2=表头)
            if col0 + 2 >= len(r):
                break
            score = to_int(r[col0])
            cnt = to_int(r[col0 + 1])
            cum = to_int(r[col0 + 2])
            if score is None:
                continue
            table.append([score, cnt or 0, cum or 0])
        one_score[year] = table

    # 等位分计算: col 6-17
    # 2026分数 2026位次 2026线差 25等位分 25等位次 24等位分 24等位次 23等位分 23等位次 22等位分 22等位次 平均等位次
    equiv = []  # list of dict
    for r in rows[2:]:   # row 3 起
        if len(r) < 17:
            break
        score_2026 = to_int(r[5])
        if score_2026 is None:
            continue
        equiv.append({
            "score26": score_2026,
            "rank26":  to_int(r[6]),
            "lineDiff26": to_int(r[7]),
            "score25": to_int(r[8]),
            "rank25":  to_int(r[9]),
            "score24": to_int(r[10]),
            "rank24":  to_int(r[11]),
            "score23": to_int(r[12]),
            "rank23":  to_int(r[13]),
            "score22": to_int(r[14]),
            "rank22":  to_int(r[15]),
            "rankAvg": to_int(r[16]),
        })

    return {
        "baseline": baseline,
        "oneScoreOneRank": one_score,
        "equivalent": equiv,
    }


# ----- 优先次序表 (来自 排名 sheet) -----
def load_priority(wb):
    ws = wb["排名"]
    rows = list(ws.iter_rows(values_only=True))

    # 城市优先次序: col 0-1 (城市, 排序)
    cities = []
    for r in rows[2:]:
        if r[0]:
            cities.append({"city": to_str(r[0]), "sort": to_int(r[1])})

    # 学校优先次序: col 3-11
    # 招生院校, 学校排名, 所在省, 城市, 院校标签, 学校类型, 主管部门, 城市排序, 学校排序
    schools = []
    for r in rows[2:]:
        if r[3]:
            schools.append({
                "name": to_str(r[3]),
                "rank": to_int(r[4]),
                "province": to_str(r[5]),
                "city": to_str(r[6]),
                "tag": to_str(r[7]),
                "type": to_str(r[8]),
                "managing": to_str(r[9]),
                "citySort": to_int(r[10]),
                "sort": to_int(r[11]),
            })

    # 专业类优先次序: col 13-15
    # 专业门类, 专业类, 专业类排序
    classes = []
    for r in rows[2:]:
        if r[13]:
            classes.append({
                "category": to_str(r[13]),
                "name": to_str(r[14]),
                "sort": to_int(r[15]),
            })

    # 专业优先次序: col 21-27
    # 专业门类, 专业类, 专业代码, 专业名称, 专业类排序, 热门排序, 专业排序
    majors = []
    for r in rows[2:]:
        if len(r) > 27 and r[24]:    # 专业名称
            majors.append({
                "category": to_str(r[21]),
                "majorClass": to_str(r[22]),
                "code": to_str(r[23]),
                "name": to_str(r[24]),
                "classSort": to_int(r[25]),
                "hotSort": to_int(r[26]),
                "sort": to_int(r[27]),
            })

    return {
        "cities": cities,
        "schools": schools,
        "majorClasses": classes,
        "majors": majors,
    }


# ----- 元数据 + 枚举 -----
def build_meta(plans):
    enums = {
        "schoolTags":      sorted({p["schoolTag"] for p in plans if p["schoolTag"]}),
        "provinces":       sorted({p["province"] for p in plans if p["province"]}),
        "cities":          sorted({p["city"] for p in plans if p["city"]}),
        "cityTiers":       sorted({p["cityTier"] for p in plans if p["cityTier"]}),
        "majorCategories": sorted({p["majorCategory"] for p in plans if p["majorCategory"]}),
        "majorClasses":    sorted({p["majorClass"] for p in plans if p["majorClass"]}),
        "schoolTypes":     sorted({p["schoolType"] for p in plans if p["schoolType"]}),
        "refConfidences":  sorted({p["refConfidence"] for p in plans if p["refConfidence"]}),
        "evalLevels":      ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-"],
    }

    cnt_total = len(plans)
    cnt_stopped = sum(1 for p in plans if p["isStopped"])
    cnt_mid = sum(1 for p in plans if p["isMidOutside"])

    return {
        "generatedAt": datetime.datetime.now().isoformat(timespec="seconds"),
        "source":      str(SRC),
        "totalPlans":  cnt_total,
        "stoppedCount": cnt_stopped,
        "nonStoppedCount": cnt_total - cnt_stopped,
        "midOutsideCount": cnt_mid,
        "enumerations": enums,
    }


def slim(obj):
    """递归去掉 None / "" / 空数组 / 空对象, 显著缩小 JSON 体积."""
    if isinstance(obj, dict):
        return {k: slim(v) for k, v in obj.items()
                if not (v is None or v == "" or v == [] or v == {})}
    if isinstance(obj, list):
        return [slim(x) for x in obj]
    return obj


# diffStructured 内的 rename 字段不能 slim (web 端模板用 .length 直接访问)
# 因此 plans 用 slim_plan 而非 slim
PRESERVE_EMPTY_LISTS = {"added", "removed", "kept"}

def slim_plan(obj, parent_key=None):
    if isinstance(obj, dict):
        return {k: slim_plan(v, k) for k, v in obj.items()
                if not (v is None or v == "" or v == [] or v == {})
                or k in PRESERVE_EMPTY_LISTS}
    if isinstance(obj, list):
        return [slim_plan(x) for x in obj]
    return obj


# ----- 各 JSON 独立生成器 (数据-1: 模块化, 支持只重生成某一项) -----
def _norm_school(s):
    """学校名归一: 全角括号 → 半角. 与 scraper 的 clean_school_name 一致."""
    if not s: return ""
    return str(s).replace("(", "(").replace(")", ")").strip()


def load_baoyan_sources():
    """加载新算法所需 3 个数据源.
    返回:
        sm2c:    dict[(校归一, 专业)] -> 学院
        sc2rate: dict[(校归一, 学院)] -> rate(float)
    """
    sm2c = {}
    if SCHOOL_MAJORS_DATA.exists():
        for f in sorted(SCHOOL_MAJORS_DATA.glob("*.csv")):
            school = _norm_school(f.stem)
            with f.open(encoding="utf-8-sig") as fh:
                for r in csv.DictReader(fh):
                    college = (r.get("学院") or "").strip()
                    major   = (r.get("专业") or "").strip()
                    if school and college and major:
                        sm2c.setdefault((school, major), college)
    # fallback: 院校-学院-专业保研率.csv 也含 (校, 学院, 专业) 三元组
    if MAJOR_BAOYAN_CSV.exists():
        with MAJOR_BAOYAN_CSV.open(encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                sch = _norm_school(r.get("学校") or "")
                col = (r.get("学院") or "").strip()
                m   = (r.get("专业") or "").strip()
                if sch and col and m:
                    sm2c.setdefault((sch, m), col)
    sc2rate = {}
    if COLLEGE_BAOYAN_CSV.exists():
        with COLLEGE_BAOYAN_CSV.open(encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                sch  = _norm_school(r.get("学校名称") or "")
                col  = (r.get("院系") or "").strip()
                rate = r.get("保研率")
                if not (sch and col and rate): continue
                try:
                    sc2rate[(sch, col)] = float(rate)
                except (ValueError, TypeError):
                    pass
    return sm2c, sc2rate


def recompute_baoyan_detail(plans):
    """V9: 用新算法重算 plan.baoyanDetail. 4 档:
       (a) 学院已知 + 学院级 rate    → 'M(学院,X%)'  [override]
       (b) 学院已知 + 仅校保研率      → 'M(学院,X%)'  [school base]
       (c) 学院未知 + 有校保研率      → 'M(本校,X%)'   [unknown college fallback]
       (d) 学院未知 + 无校保研率      → skip 该 major
       全 plan 一项都没出 → baoyanDetail = ""
    Returns: dict 统计.
    """
    sm2c, sc2rate = load_baoyan_sources()
    cnt = dict(override=0, school_base=0, unknown_college=0, skip=0, empty_plans=0)
    for p in plans:
        majors = p.get("containedMajors") or []
        school = _norm_school(p.get("schoolName") or "")
        base_rate = p.get("schoolBaoyan")    # float | None
        items = []
        for m in majors:
            college = sm2c.get((school, m))
            if college:
                override = sc2rate.get((school, college))
                if override is not None:
                    pct = round(override * 100)
                    items.append(f"{m}({college},{pct}%)")
                    cnt["override"] += 1
                elif base_rate is not None:
                    pct = round(base_rate * 100)
                    items.append(f"{m}({college},{pct}%)")
                    cnt["school_base"] += 1
                else:
                    cnt["skip"] += 1
            else:
                # 学院未知 → 用 校保研率, college 字段填"本校"以示来源
                if base_rate is not None:
                    pct = round(base_rate * 100)
                    items.append(f"{m}(本校,{pct}%)")
                    cnt["unknown_college"] += 1
                else:
                    cnt["skip"] += 1
        if items:
            p["baoyanDetail"] = ",".join(items)
        else:
            p["baoyanDetail"] = ""
            cnt["empty_plans"] += 1
    return cnt


def build_plans_json(wb, out_dir):
    print("加载招生计划 ...")
    plans = load_plans(wb)
    print(f"  共 {len(plans)} 条")
    # V9: 用新算法重算 26 专业保研率 (取代 xlsx 中预计算列)
    print("重算 26专业保研率 (新算法 4 档) ...")
    bc = recompute_baoyan_detail(plans)
    print(f"  (a) 学院已知 + 学院级 rate:     {bc['override']} 项")
    print(f"  (b) 学院已知 + 校保研率:        {bc['school_base']} 项")
    print(f"  (c) 学院未知 + 校保研率(本校):  {bc['unknown_college']} 项")
    print(f"  (d) 跳过 (校保研率也没有):      {bc['skip']} 项")
    print(f"  全空 plans (一项都没出):       {bc['empty_plans']} 条")
    plans_slim = [slim_plan(p) for p in plans]
    p = out_dir / "plans.json"
    p.write_text(json.dumps(plans_slim, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"  写入 {p} ({p.stat().st_size // 1024} KB)")
    return plans


def build_score_rank_json(wb, out_dir):
    print("加载分数 / 一分一段 / 等位分 ...")
    score_rank = load_score_rank(wb)
    print(f"  一本线 {len(score_rank['baseline'])} 年")
    print(f"  一分一段 {len(score_rank['oneScoreOneRank'])} 年")
    print(f"  等位分表 {len(score_rank['equivalent'])} 行")
    p = out_dir / "score_rank.json"
    p.write_text(json.dumps(score_rank, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"  写入 {p} ({p.stat().st_size // 1024} KB)")
    return score_rank


def build_priority_json(wb, out_dir):
    print("加载 排名 sheet 优先次序表 ...")
    priority = load_priority(wb)
    print(f"  城市 {len(priority['cities'])} 个")
    print(f"  学校 {len(priority['schools'])} 所")
    print(f"  专业类 {len(priority['majorClasses'])} 个")
    print(f"  专业 {len(priority['majors'])} 个")
    p = out_dir / "priority.json"
    p.write_text(json.dumps(priority, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"  写入 {p} ({p.stat().st_size // 1024} KB)")
    return priority


def build_meta_json(wb, out_dir, plans=None):
    """meta 依赖 plans (枚举值需要从所有行汇总)."""
    if plans is None:
        plans = load_plans(wb)
    print("生成 meta ...")
    meta = build_meta(plans)
    p = out_dir / "meta.json"
    p.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  写入 {p} ({p.stat().st_size // 1024} KB)")
    return meta


# ----- 主流程 -----
def main():
    import argparse
    ap = argparse.ArgumentParser(
        description="从 xlsx 生成 JSON. 支持 --only 单独重生成某一项."
    )
    ap.add_argument(
        "--only",
        choices=["plans", "score_rank", "priority", "meta", "all"],
        default="all",
        help="只重生成指定的一项 JSON (默认 all 全部重生成).\n"
             "  plans      - 招生计划 (15 MB, ~5s)\n"
             "  score_rank - 一分一段 + 等位分 (用户更新 26 一分一段后只需跑这个)\n"
             "  priority   - 优先次序表\n"
             "  meta       - 元数据 (依赖 plans)\n"
             "  all        - 全部",
    )
    args = ap.parse_args()

    print(f"读 {SRC}")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    only = args.only

    plans = None
    meta = None
    if only in ("all", "plans"):
        plans = build_plans_json(wb, OUT_DIR)
    if only in ("all", "score_rank"):
        build_score_rank_json(wb, OUT_DIR)
    if only in ("all", "priority"):
        build_priority_json(wb, OUT_DIR)
    if only in ("all", "meta"):
        meta = build_meta_json(wb, OUT_DIR, plans=plans)

    if only != "all":
        print(f"\n[--only {only}] 完成. 其它 JSON 未触碰.")
        return
    print(f"\n统计:")
    print(f"  总计:     {meta['totalPlans']}")
    print(f"  停招:     {meta['stoppedCount']}")
    print(f"  非停招:   {meta['nonStoppedCount']}")
    print(f"  中外合作: {meta['midOutsideCount']}")
    print(f"  学校数:   {len(meta['enumerations']['schoolTags'])} 个标签, "
          f"{len({p['schoolCode'] for p in plans})} 所学校")


if __name__ == "__main__":
    main()
