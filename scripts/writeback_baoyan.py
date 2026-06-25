#!/usr/bin/env python3
"""把 plans.json 里新算法计算的 baoyanDetail 写回 xlsx 的 2026数据raw sheet.

写入: '2026数据raw' sheet 的 '专业保研率' 列 (= 26 列).
按 '序号' 列匹配行号 (plan.id 对应 xlsx 序号).

不动 '2026数据' sheet (那里用户有自己的公式).
不动 '_2026_internal' (scraper 中转表).
"""
import json
import sys
from pathlib import Path

import openpyxl

XLSX  = Path("/Users/banlibo/Desktop/高考/new/2026高考-物理-志愿_2026.xlsx")
JSON_ = Path(__file__).resolve().parent.parent / "data" / "plans.json"
SHEET = "2026数据raw"

def main():
    print(f"读 {JSON_}")
    plans = json.loads(JSON_.read_text(encoding="utf-8"))
    by_id = {p["id"]: p.get("baoyanDetail") or "" for p in plans}
    print(f"  plans: {len(by_id)}")

    print(f"打开 {XLSX} (sheet={SHEET}) ...")
    wb = openpyxl.load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        sys.exit(f"sheet 不存在: {SHEET}")
    ws = wb[SHEET]
    hdr = [c.value for c in ws[1]]
    idx_map = {h: i for i, h in enumerate(hdr) if h is not None}
    id_col  = idx_map.get("序号")
    by_col  = idx_map.get("专业保研率")
    if id_col is None or by_col is None:
        sys.exit(f"缺列. 序号={id_col} 专业保研率={by_col}")
    print(f"  序号 col={id_col+1}, 专业保研率 col={by_col+1}")

    matched, missing = 0, 0
    for r_idx in range(2, ws.max_row + 1):
        row_id = ws.cell(r_idx, id_col + 1).value
        try:
            pid = int(row_id) if row_id is not None else None
        except (TypeError, ValueError):
            pid = None
        if pid is None:
            missing += 1
            continue
        new_val = by_id.get(pid)
        if new_val is None:
            missing += 1
            continue
        ws.cell(r_idx, by_col + 1).value = new_val or None
        matched += 1

    print(f"  匹配并写入: {matched} 行")
    print(f"  没匹配:    {missing} 行")
    print("保存 ...")
    wb.save(XLSX)
    print(f"完成: {XLSX}")

if __name__ == "__main__":
    main()
